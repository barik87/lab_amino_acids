from datetime import datetime

from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

MULTIPLIERS = dict(
    Plazma=dict(
        Alanine=325.9346088,
        Methionine=325.747619,
        Isoleucine=343.4544872,
        Lysine=187.0718163,
        Proline=724.7446809
    ),
    Urine=dict(
        Alanine=325.6844011,
        Methionine=321.8833333,
        Isoleucine=335.89375,
        Lysine=187.6297872,
        Proline=724.7446809
    )
)

CONST_MULTIPLIERS = {
    'Alanine': 2925.907,
    'Methionine': 3003.707,
    'Isoleucine': 2967.626,
    'Lysine': 940.3833,
    'Proline': 2873.455,
    'Aspartic acid': 2222.763,
    'Glutamic acid': 2174.47,
    'Serine': 2557.236,
    'Histidine': 1697.498,
    'Glycine': 2800.287,
    'Threonine': 2606.218,
    'Citruline': 2672.156,
    'Arginine': 2614.298,
    'Taurine': 3093.659,
    'Tyrosine': 2634.504,
    'Cystine': 246.4665,
    'Valine': 2895.654,
    'Phenylalanine': 2787.873,
    'Leucine': 2951.304
}

UKR_MONTHS = {
    1: 'січня',
    2: 'лютого',
    3: 'березня',
    4: 'квітня',
    5: 'травня',
    6: 'червня',
    7: 'липня',
    8: 'серпня',
    9: 'вересня',
    10: 'жовтня',
    11: 'листопада',
    12: 'грудня'
}


def get_acids(data_file_name: str, an_type: str, sample_date: datetime, birth_date: datetime) -> (str, dict,):
    diff_in_years = relativedelta(sample_date, birth_date).years
    diff_in_months = relativedelta(sample_date, birth_date).months
    if an_type == 'Plazma':
        if diff_in_years >= 12:
            return '12+ р.', get_acids_from_file(data_file_name, 13)
        if diff_in_years > 0 or diff_in_months >= 3:
            return '3м-12р', get_acids_from_file(data_file_name, 12)
        if diff_in_months < 3:
            return '0-3м.', get_acids_from_file(data_file_name, 11)
        print('ERROR: check dates!')
    else:
        if diff_in_years >= 13:
            return '13+ р.', get_acids_from_file(data_file_name, 9)
        if diff_in_years >= 7:
            return '7-13р.', get_acids_from_file(data_file_name, 8)
        if diff_in_years >= 4:
            return '4-7р.', get_acids_from_file(data_file_name, 7)
        if diff_in_years >= 2:
            return '2-4р.', get_acids_from_file(data_file_name, 6)
        if diff_in_years >= 1:
            return '1-2р.', get_acids_from_file(data_file_name, 5)
        if diff_in_months >= 6:
            return '6-12м.', get_acids_from_file(data_file_name, 4)
        if diff_in_months >= 1:
            return '1-6м.', get_acids_from_file(data_file_name, 3)
        if diff_in_months == 0:
            return '0-1м.', get_acids_from_file(data_file_name, 2)
        print('ERROR: check dates!')


def get_acids_from_file(data_file_name: str, column: int) -> dict:
    result = dict()
    wb = load_workbook(data_file_name)
    ws = wb.get_sheet_by_name('Norms')
    for row_idx in range(2, 21):
        acid_name = str(list(ws.rows)[row_idx][0].value)
        cell_values = str(list(ws.rows)[row_idx][column].value).split('--')
        result[acid_name] = (cell_values[0], cell_values[1],)
    return result


def get_acids_list(data_file_name: str) -> list:
    result = []
    wb = load_workbook(data_file_name)
    ws = wb.get_sheet_by_name('Norms')
    for row_idx in range(2, 21):
        result.append(str(list(ws.rows)[row_idx][0].value))
    return result


def get_parameters_from_file(data_file_name: str) -> dict:
    result = {}
    wb = load_workbook(data_file_name)
    ws = wb.get_sheet_by_name('Parameters')
    for row_idx in range(len(list(ws.rows))):
        row = list(ws.rows)[row_idx]
        key = row[0].value
        value = row[1].value
        if key is None or key == '':
            break
        result[key] = value
    return result
